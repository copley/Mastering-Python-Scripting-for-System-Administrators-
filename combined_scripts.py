
# --- /home/student/GithubApps/Mastering-Python-Scripting-for-System-Administrators-/Chapter01 ---

# --- 01_assign_values.py ---
name = 'John'
age = 25
address = 'USA'
percentage = 85.5
print(name)
print(age)
print(address)
print(percentage)



# AI Description:
# This code defines variables for name, age, address, and percentage. It then prints out each variable's value on a separate line.

# --- 02_access_values.py ---
#!/usr/bin/python3
str1 = 'Hello Python!'
str2 = "Object Oriented Programming"
print ("str1[0]: ", str1[0])
print ("str2[1:5]: ", str2[1:5])




# AI Description:
# This Python code defines two strings, 'Hello Python!' and 'Object Oriented Programming', then prints the first character of the first string and a slice of characters from the second string.

# --- 03_update_string.py ---
#!/usr/bin/python3
str1 = 'Hello Python!'
print("Original String: - ", str1)
print ("Updated String: - ", str1 [:6] + 'John')



# AI Description:
# This code defines a string "Hello Python!" and prints it. Then, it creates an updated string by concatenating the first 6 characters of the original string with "John" and prints the updated string.

# --- 04_string_formatting.py ---
#!/usr/bin/python3
print ("Hello this is %s and my age is %d !" % ('John', 25))



# AI Description:
# This code uses the print function to display a formatted string that includes the name "John" and the age 25. The %s and %d are placeholders for the string and integer values, respectively.

# --- 05_triple_quotes.py ---
#!/usr/bin/python3

para_str = """ Python is a scripting language which was created by
Guido van Rossum in 1991, \t which is used in various sectors such as \n Game Development, GIS Programming, Software Development, web development, 
Data Analytics and Machine learning, System Scripting etc.
"""
print (para_str)



# AI Description:
# The code defines a multi-line string containing information about Python. The string is then printed to the console, displaying details about Python's use in various sectors.

# --- 06_access_lists.py ---
#!/usr/bin/python3
cities = ['Mumbai', 'Bangalore', 'Chennai', 'Pune']
numbers = [1, 2, 3, 4, 5, 6, 7 ]
print (cities[0])
print (numbers[1:5])



# AI Description:
# This code defines two lists, 'cities' and 'numbers'. It then prints the first element of the 'cities' list and a slice of the 'numbers' list from index 1 to 4.

# --- 07_update_lists.py ---
#!/usr/bin/python3
cities = ['Mumbai', 'Bangalore', 'Chennai', 'Pune']
print ("Original Value: ", cities[3])
cities[3] = 'Delhi'
print ("New value: ", cities[3])



# AI Description:
# This code defines a list of cities and prints the original value at index 3, which is 'Pune'. It then updates the value at index 3 to 'Delhi' and prints the new value, which is 'Delhi'.

# --- 08_delete_lists.py ---
#!/usr/bin/python3
cities = ['Mumbai', 'Bangalore', 'Chennai', 'Pune']
print ("Before deleting: ", cities)
del cities[2]
print ("After deleting: ", cities)



# AI Description:
# This code initializes a list of cities, prints the list, deletes the city at index 2 (Chennai), and then prints the updated list without the deleted city.

# --- 09_access_tuples.py ---
#!/usr/bin/python3
cities = ('Mumbai', 'Bangalore', 'Chennai', 'Pune')
numbers = (1, 2, 3, 4, 5, 6, 7)
print (cities[3])
print (numbers[1:6])


# AI Description:
# This code defines two tuples: 'cities' with city names and 'numbers' with integers. It then prints the element at index 3 from the 'cities' tuple and a slice of elements from index 1 to 5 from the 'numbers' tuple.

# --- 10_update_tuples.py ---
#!/usr/bin/python3
cities = ('Mumbai', 'Bangalore', 'Chennai', 'Pune')
numbers = (1,2,3,4,5,6,7)
tuple1 = cities + numbers
print(tuple1)



# AI Description:
# This code creates two tuples, one with city names and another with numbers. It then combines the two tuples into a new tuple and prints the result.

# --- 11_delete_tuples.py ---
#!/usr/bin/python3
cities = ('Mumbai', 'Bangalore', 'Chennai', 'Pune')
print ("Before deleting: ", cities)
del cities
print ("After deleting: ", cities)



# AI Description:
# This code defines a tuple named 'cities' containing city names. It then deletes the tuple using the 'del' keyword and attempts to print the tuple after deletion, which results in a NameError as the tuple no longer exists.

# --- 12_argv_example.py ---
import sys
print('Number of arguments:', len(sys.argv))
print('Argument list:', str(sys.argv))



# AI Description:
# This code imports the sys module and prints the number of command-line arguments passed when running the script, as well as the list of those arguments.

# --- 13_if_example.py ---
a = 10
if a > 0:
	print(a, "is a positive number.")
print("This statement is always printed.")
a = -10
if a > 0:
	print(a, "is a positive number.")



# AI Description:
# The code sets variable 'a' to 10 and checks if it's greater than 0, then prints a message if it's positive. The statement "This statement is always printed." is printed regardless of the value of 'a'. Variable 'a' is then set to -10 and the code skips the second if statement because -10 is not greater than 0.

# --- 14_if_else_example.py ---
a = 10
if a > 0:
	print("Positive number")
else:
	print("Negative number")



# AI Description:
# The code defines a variable `a` with the value 10. It then checks if `a` is greater than 0 and prints "Positive number" if it is, otherwise it prints "Negative number".

# --- 15_if_elif_example.py ---
a = 10
if a > 50:
    print("a is greater than 50")
elif a == 10:
    print("a is equal to 10")
else:
    print("a is negative")



# AI Description:
# The code sets the variable 'a' to 10. It then checks if 'a' is greater than 50, equal to 10, or negative, and prints a corresponding message based on the condition. In this case, it will print "a is equal to 10".

# --- 16_for_example.py ---
numbers = [6, 5, 3, 8, 4, 2, 5, 4, 11]
sum = 0
for i in numbers:
	sum = sum + i
print("The sum is", sum)



# AI Description:
# This code initializes a list of numbers, calculates the sum of all the numbers in the list using a loop, and then prints the final sum.

# --- 17_range_example.py ---
for i in range(5):
	print("The number is", i)



# AI Description:
# This code uses a for loop to iterate through numbers 0 to 4. It prints a message along with the current number during each iteration.

# --- 18_while_example.py ---
a = 10
sum = 0
i = 1
while i <= a:
	sum = sum + i
	i = i + 1
print("The sum is", sum)



# AI Description:
# This code initializes variables a, sum, and i. It then calculates the sum of integers from 1 to a using a while loop. Finally, it prints the sum.

# --- 19_iterator_example.py ---
numbers = [10, 20, 30, 40]
numbers_iter = iter(numbers)
print(next(numbers_iter))
print(next(numbers_iter))
print(numbers_iter.__next__())
print(numbers_iter.__next__())
next(numbers_iter)


# AI Description:
# The code creates a list of numbers and an iterator object from the list. It then prints the next element from the iterator using both the `next()` function and the `__next__()` method. Finally, it attempts to call `next()` on the iterator object, which will raise a `StopIteration` error as there are no more elements to iterate over.

# --- 20_generator_example.py ---
def my_gen():
	n = 1
	print('This is printed first')
	yield n
	n += 1
	print('This is printed second')
	yield n
	n += 1
	print('This is printed at last')
	yield n
for item in my_gen():
	print(item) 



# AI Description:
# This code defines a generator function called my_gen that yields numbers incrementally and prints a message each time. The generator is iterated over using a for loop to print each yielded number and message sequentially.

# --- 21_function_example.py ---
def welcome(name):
	print("Hello " + name + ", Welcome to Python Programming !")
welcome("John")



# AI Description:
# This code defines a function called "welcome" that takes a name as input and prints a greeting message. It then calls the function with the name "John", resulting in the output: "Hello John, Welcome to Python Programming!"

# --- 22_return_statement.py ---
def return_value(a):
	if a >= 0:
		return a
	else:
		return -a
print(return_value(2))
print(return_value(-4))



# AI Description:
# This code defines a function called `return_value` that returns the absolute value of the input parameter `a`. It then calls the function twice with different arguments (2 and -4) and prints the results. The output will be 2 and 4, respectively.

# --- 23_lambda_using_filter.py ---
numbers = [10, 25, 54, 86, 89, 11, 33, 22]
new_numbers  = list(filter(lambda x: (x%2 == 0) , numbers))
print(new_numbers)



# AI Description:
# This code creates a list of numbers. It then filters the list to keep only the even numbers using a lambda function. Finally, it prints the new list of even numbers.

# --- 24_lambda_using_map.py ---
my_list = [1, 5, 4, 6, 8, 11, 3, 12]
new_list = list(map(lambda x: x * 2 , my_list))
print(new_list)



# AI Description:
# The code creates a list called `my_list`. It then uses a lambda function with `map` to double each element in `my_list` and stores the results in a new list called `new_list`. Finally, it prints the `new_list` containing the doubled values of the original list.

# --- add.py ---
import sample
sum = sample.addition(10, 20)
print(sum)



# AI Description:
# This code imports a module named 'sample' and calls a function 'addition' from this module with arguments 10 and 20. The result of the function call is stored in a variable 'sum' and then printed to the console.

# --- sample.py ---
def addition(num1, num2):
	result = num1 + num2
	return result



# AI Description:
# This Python code defines a function called "addition" that takes two input parameters, num1 and num2, and returns their sum.


# --- /home/student/GithubApps/Mastering-Python-Scripting-for-System-Administrators-/Chapter02 ---

# --- cprof_example.py ---
mul_value = 0 
def mul_numbers( num1, num2 ):
	mul_value = num1 * num2; 
	print ("Local Value: ", mul_value)
	return mul_value
mul_numbers( 58, 77 )
print ("Global Value: ", mul_value)



# AI Description:
# This code defines a function 'mul_numbers' that calculates the product of two numbers and stores it in a local variable 'mul_value'. The function then prints and returns the local value. The global variable 'mul_value' remains 0, as it is not updated within the function.

# --- exception_example.py ---
a = 35
b = 57
try:
	c = a + b
	print("The value of c is: ", c)
	d = b / 0
	print("The value of d is: ", d)

except:
	print("Division by zero is not possible")

print("Out of try...except block")




# AI Description:
# The code initializes variables 'a' and 'b' to 35 and 57, respectively. It then performs addition and division operations on these variables within a try-except block to handle division by zero, printing the results or an error message accordingly. Finally, it prints "Out of try...except block" regardless of the exception being caught.

# --- pdb_example.py ---
class Student:
	def __init__(self, std):
		self.count = std

	def print_std(self):
		for i in range(self.count):
			print(i)
		return
if __name__ == '__main__':
	Student(5).print_std()



# AI Description:
# This code defines a class called Student with an initializer that sets a count attribute. The class also has a method to print numbers up to the count value. When the code is run, it creates an instance of the Student class with a count of 5 and calls the print_std method to print numbers from 0 to 4.

# --- timeit_example.py ---
import timeit
prg_setup = "from math import sqrt"
prg_code = '''
def timeit_example():
	list1 = []
	for x in range(50):
		list1.append(sqrt(x))
''' 
# timeit statement
print(timeit.timeit(setup = prg_setup, stmt = prg_code, number = 10000))



# AI Description:
# This code imports the timeit module and defines a setup statement importing the sqrt function from the math module. It then defines a function that appends the square root of numbers to a list, and uses timeit to measure the execution time of this function being called 10000 times. The final result is printed out.

# --- trace_example.py ---
class Student:
	def __init__(self, std):
		self.count = std

	def go(self):
		for i in range(self.count):
			print(i)
		return
if __name__ == '__main__':
	Student(5).go()


# AI Description:
# This code defines a class Student with an initialization method that sets a count attribute. The class has a method go() that prints numbers from 0 to count-1. Finally, the code creates an instance of Student with count as 5 and calls the go() method.


# --- /home/student/GithubApps/Mastering-Python-Scripting-for-System-Administrators-/Chapter03 ---

# --- arithmetic.py ---
# In this script, we are going to create a 4 functions: add_numbers, sub_numbers, mul_numbers, div_numbers.
def add_numbers(x, y):
	return x + y

def sub_numbers(x, y):
	return x - y

def mul_numbers(x, y):
	return x * y

def div_numbers(x, y):
	return (x / y)



# AI Description:
# This Python script defines four functions: add_numbers, sub_numbers, mul_numbers, and div_numbers to perform basic arithmetic operations on two input numbers. The functions return the sum, difference, product, and division of the input numbers, respectively.

# --- if_example.py ---
def check_if():
	a = int(input("Enter a number \n"))
	if (a == 100):
		print("a is equal to 100")

	else:
		print("a is not equal to 100")

	return a


# AI Description:
# This code defines a function that prompts the user to enter a number and checks if it is equal to 100. If the number is 100, it prints a message stating that. Otherwise, it prints a message indicating that the number is not equal to 100. The function returns the entered number.

# --- test_arithmetic.py ---
import arithmetic
import unittest
# Testing add_numbers function from arithmetic.
class Test_addition(unittest.TestCase):
	# Testing Integers
	def test_add_numbers_int(self):
		sum = arithmetic.add_numbers(50, 50)
		self.assertEqual(sum, 100)

	# Testing Floats
	def test_add_numbers_float(self):
		sum = arithmetic.add_numbers(50.55, 78)
		self.assertEqual(sum, 128.55)

	# Testing Strings
	def test_add_numbers_strings(self):
		sum = arithmetic.add_numbers('hello','python')
		self.assertEqual(sum, 'hellopython')

class Test_subtraction(unittest.TestCase):
	# Testing Integers
	def test_sub_numbers_int(self):
		diff = arithmetic.sub_numbers(50, 50)
		self.assertEqual(diff, 0)

	# Testing Floats
	def test_sub_numbers_float(self):
		diff = arithmetic.sub_numbers(80.00, 78)
		self.assertEqual(diff, 2.00)

class Test_multiplication(unittest.TestCase):
	# Testing Integers
	def test_mul_numbers_int(self):
		multi = arithmetic.mul_numbers(78, 46)
		self.assertEqual(multi, 3588)

	# Testing Floats
	def test_mul_numbers_float(self):
		multi = arithmetic.mul_numbers(77.85, 8)
		self.assertEqual(multi, 622.8)

class Test_division(unittest.TestCase):
	# Testing Integers
	def test_div_numbers_int(self):
		quotient = arithmetic.div_numbers(78, 2)
		self.assertEqual(quotient, 39)

	# Testing Floats
	def test_div_numbers_float(self):
		quotient = arithmetic.div_numbers(77.8, 2)
		self.assertEqual(quotient, 38.9)

if __name__ == '__main__':
	unittest.main()


# AI Description:
# This Python code includes unit tests for functions such as addition, subtraction, multiplication, and division in the 'arithmetic' module. The tests cover scenarios with integers, floats, and strings as inputs. The unittest library is used to run the tests when the script is executed.

# --- test_if.py ---
import if_example
import unittest

class Test_if(unittest.TestCase):
	def test_if(self):
		result = if_example.check_if()
		self.assertEqual(result, 100)
		
if __name__ == '__main__':
	unittest.main()



# AI Description:
# This code imports a module named if_example and unittest. It defines a test case class Test_if that checks the return value of a function check_if from if_example to be 100. Finally, it runs the unit tests using unittest.main().


# --- /home/student/GithubApps/Mastering-Python-Scripting-for-System-Administrators-/Chapter04 ---

# --- accept_by_input_file.py ---
i = open('sample.txt','r') 
o = open('sample_output.txt','w') 
a = i.read() 
o.write(a)



# AI Description:
# This Python code opens a file named 'sample.txt' in read mode and another file named 'sample_output.txt' in write mode. It then reads the content from 'sample.txt' and writes it to 'sample_output.txt'.

# --- accept_by_pipe.py ---
import sys
for n in sys.stdin:
	print ( int(n.strip())//2 )



# AI Description:
# This code reads input from the console using sys.stdin and prints half of each input number after stripping any whitespace characters.

# --- capture_output.py ---
import subprocess
res = subprocess.run(['ls', '-1'], stdout=subprocess.PIPE,)
print('returncode:', res.returncode)
print(' {} bytes in stdout:\n{}'.format(len(res.stdout), res.stdout.decode('utf-8')))



# AI Description:
# This code uses the subprocess module to run the 'ls -1' command. It captures the output and decodes it to a UTF-8 string. Finally, it prints the return code, the length of the output, and the decoded output.

# --- execute_external_commands.py ---
import subprocess

subprocess.call(["touch", "sample.txt"])
subprocess.call(["ls"])
print("Sample file created")
subprocess.call(["rm", "sample.txt"])
subprocess.call(["ls"])
print("Sample file deleted")



# AI Description:
# The code uses the subprocess module to create a file named "sample.txt" using the touch command and then lists the files in the directory using the ls command. It prints a message indicating that the sample file is created, then deletes the file using the rm command and lists the files again before printing a message indicating that the sample file is deleted.

# --- generate_warnings.py ---
import warnings
warnings.simplefilter('error', UserWarning)
print('Before')
warnings.warn('Write your warning message here')
print('After')



# AI Description:
# This code sets a filter to raise an error for UserWarning. It then prints 'Before', raises a UserWarning with a custom message, and finally prints 'After'.

# --- getpass_example.py ---
import getpass
passwd = getpass.getpass(prompt='Enter your password: ')
if passwd.lower() == '#pythonworld':
	print('Welcome!!')
else:
	print('The password entered is incorrect!!')



# AI Description:
# This code uses the `getpass` module to prompt the user to enter a password securely. If the entered password is '#pythonworld' (case-insensitive), it prints 'Welcome!!', otherwise it prints 'The password entered is incorrect!!'.

# --- handling_password.py ---
import sys
import paramiko
import time
ip_address = "192.168.2.106"
username = "student"
password = "training"
ssh_client = paramiko.SSHClient()
ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
ssh_client.load_system_host_keys()
ssh_client.connect(hostname=ip_address,\
			username=username, password=password)
print ("Successful connection", ip_address)
ssh_client.invoke_shell()
remote_connection = ssh_client.exec_command('cd Desktop; mkdir work\n')
remote_connection = ssh_client.exec_command('mkdir test_folder\n')
#print( remote_connection.read() )
ssh_client.close



# AI Description:
# This code connects to a remote server using SSH, creates a new directory named 'work' in the Desktop folder, and another directory named 'test_folder'. The code then closes the SSH connection.

# --- list_dir.py ---
import os
import sys
print(sorted(os.listdir(sys.argv[1])))



# AI Description:
# This code imports necessary modules, then prints a sorted list of files and directories in the directory path specified as a command line argument.

# --- logging_example.py ---
import logging
LOG_FILENAME = 'hello.py'
logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
logging.debug('This message should go to the log file')
with open(LOG_FILENAME, 'rt') as f:
	prg = f.read()
print('FILE:')
print(prg) 



# AI Description:
# This code sets up logging to a file named 'hello.py' and logs a debug message. It then reads the contents of the log file and prints it to the console.

# --- logging_warnings_codes.py ---
import logging
import warnings

logging.basicConfig(level=logging.INFO,)
warnings.warn('This warning is not sent to the logs')
logging.captureWarnings(True)
warnings.warn('This warning is sent to the logs')



# AI Description:
# The code sets up logging with the INFO level. It then issues a warning that is not sent to the logs. After enabling capturing warnings to logs, another warning is issued and this one is sent to the logs.

# --- no_prompt.py ---
import getpass
try:
	p = getpass.getpass()
except Exception as error:
	print('ERROR', error)
else:
	print('Password entered:', p)



# AI Description:
# This code imports the getpass module to securely input a password. It prompts the user to enter a password, catches any exceptions that occur, and prints the entered password if there are no errors.

# --- open_web.py ---
import webbrowser
webbrowser.open('https://timesofindia.indiatimes.com/world')



# AI Description:
# This code imports the webbrowser module and uses it to open a web browser to the specified URL 'https://timesofindia.indiatimes.com/world'.

# --- os_dir_example.py ---
import os
directory_name = 'abcd'
print('Creating', directory_name)
os.makedirs(directory_name)
file_name = os.path.join(directory_name, 'sample_example.txt')
print('Creating', file_name)
with open(file_name, 'wt') as f:
	f.write('sample example file')
print('Cleaning up')
os.unlink(file_name)
os.rmdir(directory_name)	# Will delete the directory



# AI Description:
# This code creates a directory named 'abcd', then creates a file named 'sample_example.txt' inside that directory with some text. It then deletes the file and the directory.

# --- password_prompt_again.py ---
import getpass
user_name = getpass.getuser()
print ("User Name : %s" % user_name)
while True:
	passwd = getpass.getpass("Enter your Password : ")
 	if passwd == '#pythonworld':
		print ("Welcome!!!")
		break
	else:
		print ("The password you entered is incorrect.")



# AI Description:
# This code imports the getpass module to prompt the user for their username and password. It then checks if the entered password is '#pythonworld' and displays a welcome message if it matches, otherwise, it prompts the user to re-enter the password.

# --- put_cpu_limit.py ---
import resource
import sys
import signal
import time
def time_expired(n, stack):
	print('EXPIRED :', time.ctime())
	raise SystemExit('(time ran out)')
signal.signal(signal.SIGXCPU, time_expired)
# Adjust the CPU time limit
soft, hard = resource.getrlimit(resource.RLIMIT_CPU)
print('Soft limit starts as  :', soft)
resource.setrlimit(resource.RLIMIT_CPU, (10, hard))
soft, hard = resource.getrlimit(resource.RLIMIT_CPU)
print('Soft limit changed to :', soft)
print()
# Consume some CPU time in a pointless exercise
print('Starting:', time.ctime())
for i in range(200000):
	for i in range(200000):
		v = i * i
# We should never make it this far
print('Exiting :', time.ctime())



# AI Description:
# The code sets a CPU time limit to 10 seconds using the `resource` module. It then consumes CPU time in a loop and raises a SystemExit if the time limit is exceeded. The code demonstrates setting and enforcing a CPU time limit for a process in Python.

# --- read_config_file.py ---
from configparser import ConfigParser
p = ConfigParser()
p.read('read_simple.ini')
print(p.get('bug_tracker', 'url'))



# AI Description:
# This code reads a configuration file named 'read_simple.ini' using ConfigParser, retrieves the value associated with the key 'url' under the section 'bug_tracker', and then prints this value.

# --- read_many_config_file.py ---
from configparser import ConfigParser
import glob
p = ConfigParser()
files = ['hello.ini', 'bye.ini', 'read_simple.ini', 'welcome.ini']
files_found = p.read(files)
files_missing = set(files) - set(files_found)
print('Files found:  ', sorted(files_found))
print('Files missing:  ', sorted(files_missing))



# AI Description:
# This code imports ConfigParser and glob modules, creates a ConfigParser object, and attempts to read configuration files specified in a list. It then calculates the files found and missing from the list, and prints the sorted lists of found and missing files.

# --- redirection.py ---
import sys
class Redirection(object):
	def __init__(self, in_obj, out_obj):
		self.input = in_obj
		self.output = out_obj

	def read_line(self):
		res = self.input.readline()
		self.output.write(res)
		return res

if __name__ == '__main__':
	if not sys.stdin.isatty():
		sys.stdin = Redirection(in_obj=sys.stdin, out_obj=sys.stdout)

	a = input('Enter a string: ')
	b = input('Enter another string: ')
	print ('Entered strings are: ', repr(a), 'and', repr(b))



# AI Description:
# This Python code defines a class called Redirection for redirecting input and output. If the standard input is not a terminal, it redirects input to stdin and output to stdout. It then prompts the user to enter two strings and prints the entered strings along with their representations.

# --- take_backup.py ---
import os
import shutil
import time
from sh import rsync
def check_dir(os_dir):
	if not os.path.exists(os_dir):
		print (os_dir, "does not exist.")
		exit(1)
def ask_for_confirm():
	ans = input("Do you want to Continue? yes/no\n")
	global con_exit
	if ans == 'yes':
		con_exit = 0
		return con_exit
	elif ans == "no":
		con_exit = 1
		return con_exit
	else:
		print ("Answer with yes or no.")
		ask_for_confirm()
def delete_files(ending):
	for r, d, f in os.walk(backup_dir):
		for files in f:
			if files.endswith("." + ending):
				os.remove(os.path.join(r, files))
backup_dir = input("Enter directory to backup\n")	# Enter directory name
check_dir(backup_dir)
print (backup_dir, "saved.")
time.sleep(3)
backup_to_dir= input("Where to backup?\n")
check_dir(backup_to_dir)
print ("Doing the backup now!")
ask_for_confirm()
if con_exit == 1:
	print ("Aborting the backup process!")
	exit(1)
rsync("-auhv", "--delete", "--exclude=lost+found", "--exclude=/sys", "--exclude=/tmp", "--exclude=/proc",
  "--exclude=/mnt", "--exclude=/dev", "--exclude=/backup", backup_dir, backup_to_dir)



# AI Description:
# This Python code defines functions to check if a directory exists, confirm user input, and delete files with a specific file extension. It then prompts the user to enter directories for backup and destination, confirms the backup action, and uses rsync to perform the backup with specific exclusions.

# --- with_prompt.py ---
import getpass
try:
	p = getpass.getpass("Enter your password: ")
except Exception as error:
	print('ERROR', error)
else:
	print('Password entered:', p)



# AI Description:
# This code imports the getpass module to securely prompt the user for a password. It then tries to get the password input from the user and prints an error message if an exception occurs. If successful, it prints the password entered by the user.


# --- /home/student/GithubApps/Mastering-Python-Scripting-for-System-Administrators-/Chapter05 ---

# --- compare_data.py ---
import pandas as pd

df1 = pd.read_csv("student1.csv")
df2 = pd.read_csv("student2.csv")

s1 = set([ tuple(values) for values in df1.values.tolist()])
s2 = set([ tuple(values) for values in df2.values.tolist()])

s1.symmetric_difference(s2)

print (pd.DataFrame(list(s1.difference(s2))),'\n\n')
print (pd.DataFrame(list(s2.difference(s1))),'\n\n')


# AI Description:
# This code reads data from two CSV files into pandas DataFrames, converts the DataFrames into sets of tuples, and finds the symmetric difference between the two sets. It then prints the data that is present in one set but not the other set in tabular form using pandas DataFrames.

# --- compress_a_directory.py ---
from shutil import make_archive
import os
archive_name = os.path.expanduser(os.path.join('~', 'work1'))
root_dir = os.path.expanduser(os.path.join('~', '.ssh'))
make_archive(archive_name, 'gztar', root_dir)



# AI Description:
# This Python code creates a gzipped tar archive of the '.ssh' directory in the user's home directory and saves it as 'work1.tar.gz'. It uses the shutil module's make_archive function for creating the archive.

# --- examine_tar_file_content.py ---
import tarfile

tar_file = tarfile.open("work.tar.gz", "r:gz")
print(tar_file.getnames())



# AI Description:
# This code imports the `tarfile` module and opens a `.tar.gz` file named "work.tar.gz" in read mode. It then prints the names of all files and directories within the archive.

# --- hello.py ---
print ("")
print ("Hello World\n")
print ("Hello Python\n")


# AI Description:
# This code consists of three print statements that display empty space, "Hello World", and "Hello Python" respectively. The "\n" character is used to create a new line for each string.

# --- merge_data.py ---
import pandas as pd

df1 = pd.read_csv("student1.csv")
df2 = pd.read_csv("student2.csv")

result = pd.concat([df1, df2])
print(result)


# AI Description:
# The code imports the pandas library and reads data from two CSV files, "student1.csv" and "student2.csv", into separate dataframes. It then concatenates the two dataframes into a single dataframe called 'result' and prints the combined data.

# --- metadata_example.py ---
import pyPdf
def main():
	file_name = '/home/student/sample_pdf.pdf'
	pdfFile = pyPdf.PdfFileReader(file(file_name,'rb'))
	pdf_data = pdfFile.getDocumentInfo()
	print ("----Metadata of the file----")
	
	for md in pdf_data:
		print (md+ ":" +pdf_data[md])
if __name__ == '__main__':
	main()



# AI Description:
# This code imports the pyPdf library and reads metadata from a PDF file 'sample_pdf.pdf'. It then prints out the metadata keys and values.

# --- os_remove_file_directory.py ---
import os

os.remove('sample.txt')
print("File removed successfully")

os.rmdir('work1')
print("Directory removed successfully")


# AI Description:
# This code uses the 'os' module to remove a file named 'sample.txt' and print a success message. It then removes a directory named 'work1' and prints another success message.

# --- pattern_match.py ---
import glob

file_match = glob.glob('*.txt')
print(file_match)

file_match = glob.glob('[0-9].txt')
print(file_match)

file_match = glob.glob('**/*.txt', recursive=True)
print(file_match)

file_match = glob.glob('**/', recursive=True)
print(file_match)



# AI Description:
# The code uses the `glob` module to find files matching certain patterns in the current directory. The first `glob.glob('*.txt')` searches for all files with a '.txt' extension. The next one `glob.glob('[0-9].txt')` looks for files with names starting with a digit and ending with '.txt'. The third one `glob.glob('**/*.txt', recursive=True)` searches for '.txt' files in all subdirectories recursively, while the last one `

# --- shutil_copy_example.py ---
import shutil
import os

shutil.copyfile('hello.py', 'welcome.py')
print("Copy Successful\n")


# AI Description:
# This code imports the shutil and os modules, copies a file named 'hello.py' to 'welcome.py', and then prints "Copy Successful" to the console.

# --- shutil_move_example.py ---
import shutil

shutil.move('/home/student/work/sample.txt', '/home/student/Desktop')


# AI Description:
# This code uses the shutil module to move a file called sample.txt from the '/home/student/work' directory to the '/home/student/Desktop' directory.

# --- shutil_rename_example.py ---
import shutil
shutil.move('sample.bin', 'sample.txt')


# AI Description:
# The code uses the shutil module to move a file named 'sample.bin' to a new location and rename it to 'sample.txt'.

# --- tarfile_example.py ---
import tarfile

tar_file = tarfile.open("work.tar.gz", "w:gz")
for name in ["welcome.py", "hello.py", "hello.txt", "sample.txt", "sample1.txt"]:
	tar_file.add(name)
tar_file.close()


# AI Description:
# This code creates a new gzipped tar file named "work.tar.gz" and adds multiple files ("welcome.py", "hello.py", "hello.txt", "sample.txt", "sample1.txt") to it. Finally, it closes the tar file.

# --- unzip_a_directory.py ---
import shutil
shutil.unpack_archive('work1.zip')



# AI Description:
# This code imports the "shutil" module and then unpacks the contents of a ZIP file named "work1.zip" in the current directory.


# --- /home/student/GithubApps/Mastering-Python-Scripting-for-System-Administrators-/Chapter06 ---

# --- add_to_archive.py ---
import shutil
import os
import tarfile

print('creating archive')
shutil.make_archive('work', 'tar', root_dir='..', base_dir='work',)

print('\nArchive contents:')
with tarfile.open('work.tar', 'r') as t_file:
	for names in t_file.getnames():
		print(names)

os.system('touch sample.txt')
print('adding sample.txt')
with tarfile.open('work.tar', mode='a') as t:
	t.add('sample.txt')

print('contents:',)
with tarfile.open('work.tar', mode='r') as t:
	print([m.name for m in t.getmembers()])


# AI Description:
# This code creates a tar archive 'work.tar' containing files from the 'work' directory, displays its contents, adds a new file 'sample.txt' to the archive, and then prints the updated contents of the archive.

# --- check_archive_file.py ---
import tarfile

for f_name in ['hello.py', 'work.tar.gz', 'welcome.py', 'nofile.tar', 'sample.tar.xz']:
	try:
		print('{:}	{}'.format(f_name, tarfile.is_tarfile(f_name)))
	except IOError as err:
		print('{:}	{}'.format(f_name, err))


# AI Description:
# This code imports the tarfile module and checks if the given files are valid tar archives using the is_tarfile function. It loops through a list of file names and prints whether each file is a tar archive or not. If there's an IOError, it prints the file name along with the error message.

# --- check_zip_file.py ---
import zipfile

for f_name in ['hello.py', 'work.zip', 'welcome.py', 'sample.txt', 'test.zip']:
	try:
		print('{:}	{}'.format(f_name, zipfile.is_zipfile(f_name)))
	except IOError as err:
		print('{:}	{}'.format(f_name, err))



# AI Description:
# This code imports the `zipfile` module and iterates over a list of file names. It prints whether each file is a zip file using the `zipfile.is_zipfile()` function. If there is an IOError, it prints the error message.

# --- extract_contents.py ---
import tarfile
import os

os.mkdir('work')
with tarfile.open('work.tar', 'r') as t:
	t.extractall('work')
print(os.listdir('work'))


# AI Description:
# This code creates a new directory called 'work' using os.mkdir(). It then opens a tar file 'work.tar' and extracts all its contents into the 'work' directory using tarfile.open(). Finally, it prints the list of files in the 'work' directory using os.listdir().

# --- file_decrypt.py ---
import pyAesCrypt
from os import stat, remove

bufferSize = 64 * 1024
password = "#Training"

encFileSize = stat("sample.txt.aes").st_size

with open("sample.txt.aes", "rb") as fIn:
	with open("sampleout.txt", "wb") as fOut:
		try:
			pyAesCrypt.decryptStream(fIn, fOut, password, bufferSize, encFileSize)
		except ValueError:
			remove("sampleout.txt")


# AI Description:
# This code imports necessary modules and defines a buffer size and password for AES encryption. It decrypts a file "sample.txt.aes" using the specified password and writes the decrypted content to a new file "sampleout.txt". If decryption fails, it removes the newly created file.

# --- file_encrypt.py ---
import pyAesCrypt
from os import stat, remove
# encryption/decryption buffer size - 64K
bufferSize = 64 * 1024
password = "#Training"

# encrypt
with open("sample.txt", "rb") as fIn:
	with open("sample.txt.aes", "wb") as fOut:
		pyAesCrypt.encryptStream(fIn, fOut, password, bufferSize)

# get encrypted file size
encFileSize = stat("sample.txt.aes").st_size


# AI Description:
# This code imports necessary modules for encryption, sets a buffer size and password, then encrypts a file named "sample.txt" into "sample.txt.aes" using AES encryption with the specified password. Finally, it retrieves the size of the encrypted file.

# --- make_zip_file.py ---
import shutil
shutil.make_archive('work', 'zip', 'work')


# AI Description:
# This code uses the shutil module to create a zip archive of the 'work' directory. The make_archive function takes the name of the archive, the archive format (in this case, 'zip'), and the directory to archive.

# --- read_metadata.py ---
import tarfile
import time
with tarfile.open('work.tar', 'r') as t:
	for file_info in t.getmembers():
		print(file_info.name)
		print("Size   :", file_info.size, 'bytes')
		print("Type   :", file_info.type)
		print()



# AI Description:
# This code opens a TAR file named 'work.tar' in read mode and iterates over its members. For each member, it prints the file name, size in bytes, and type.

# --- read_metadata1.py ---
import zipfile
def meta_info(names):
	with zipfile.ZipFile(names) as zf:
		for info in zf.infolist():
			print(info.filename)
			if info.create_system == 0:
				system = 'Windows'
			elif info.create_system == 3:
				system = 'Unix'
			else:
				system = 'UNKNOWN'
			print("System         :", system)
			print("Zip Version    :", info.create_version)
			print("Compressed     :", info.compress_size, 'bytes')
			print("Uncompressed   :", info.file_size, 'bytes')
			print()


if __name__ == '__main__':
	meta_info('work.zip')



# AI Description:
# This Python code defines a function `meta_info` that extracts metadata information from a ZIP file such as the system it was created on, version, compressed and uncompressed sizes of files within the ZIP. The function takes a file name as input, opens it as a ZIP file, and iterates through the file info to extract and print the metadata. When executed, it calls `meta_info` function with the file name 'work.zip'.

# --- shutil_make_archive.py ---
import tarfile
import shutil
import sys

shutil.make_archive(
	'work', 'gztar',
	root_dir='..',
	base_dir='work',
)

print('\nArchive contents:')
with tarfile.open('work.tar.gz', 'r') as t_file:
	for names in t_file.getnames():
		print(names)


# AI Description:
# This code creates a gzipped tar archive named 'work.tar.gz' containing files from the 'work' directory located one level up. It then prints the contents of the archive by opening and reading it using the `tarfile` module.

# --- shutil_unpack_archive.py ---
import pathlib
import shutil
import sys
import tempfile

with tempfile.TemporaryDirectory() as d:
	shutil.unpack_archive('work.tar.gz', extract_dir='/home/student/work',)
	prefix_len = len(d) + 1
	for extracted in pathlib.Path(d).rglob('*'):
		print(str(extracted)[prefix_len:])


# AI Description:
# This code imports necessary modules, creates a temporary directory, extracts a tar.gz file into a specified directory, determines the prefix length for the extracted files, and prints the relative paths of all extracted files in the temporary directory.


# --- /home/student/GithubApps/Mastering-Python-Scripting-for-System-Administrators-/Chapter07 ---

# --- dedent_example.py ---
import textwrap

str1 = '''
	Hello Python World \tThis is Python 101
	Scripting language\n
	Python is an interpreted high-level programming language for general-purpose programming.
	'''
print("Original: \n", str1)
print()
t = textwrap.dedent(str1)
print("Dedented: \n", t)


# AI Description:
# The code imports the textwrap module and assigns a multi-line string to a variable. It then prints the original string and the dedented version of the string using textwrap.dedent(). The dedent function removes common leading whitespace from each line in the string.

# --- fill_example.py ---
import textwrap

sample_string = '''Python is an interpreted high-level programming language.'''

w = textwrap.fill(text=sample_string, width=50) 
  
print(w)


# AI Description:
# This code imports the textwrap module and creates a sample string. It then uses the textwrap.fill method to wrap the text to a specified width of 50 characters and prints the wrapped text.

# --- indent_example.py ---
import textwrap

str1 = "Python is an interpreted high-level programming language for general-purpose programming. Created by Guido van Rossum and first released in 1991, \n\nPython has a design philosophy that emphasizes code readability, notably using significant whitespace."

w = textwrap.fill(str1, width=30)
i = textwrap.indent(w, '*')
print(i)


# AI Description:
# This code imports the textwrap module to format a long string into multiple lines with a width of 30 characters. It then indents the formatted text with asterisks and prints the indented text.

# --- re_findall_example.py ---
import re

pattern = 'Red'
colors = 'Red, Blue, Black, Red, Green'
p = re.findall(pattern, colors)
print(p)

str_line = 'Peter Piper picked a peck of pickled peppers. How many pickled peppers did Peter Piper pick?'
pt = re.findall('pe\w+', str_line)
pt1 = re.findall('pic\w+', str_line)
print(pt)
print(pt1)

line = 'Hello hello HELLO bye'
p = re.findall('he\w+', line, re.IGNORECASE)
print(p)




# AI Description:
# The code uses the re module to find all occurrences of the string 'Red' in a list of colors and prints the result. It then finds words starting with 'pe' and 'pic' in a given string and prints the matches. Lastly, it searches for words starting with 'he' in a case-insensitive manner in a string and prints the matches.

# --- re_match.py ---
import re

str_line = "This is python tutorial. Do you enjoy learning python ?"

obj = re.match(r'(.*) enjoy (.*?) .*', str_line)

if obj:
	print(obj.groups())


# AI Description:
# This code imports the 're' module for regular expressions in Python. It uses a regex pattern to search for a specific phrase in the 'str_line'. If a match is found, it prints the groups captured by the regex pattern.

# --- re_search.py ---
import re

pattern = ['programming', 'hello']
str_line = 'Python programming is fun'

for p in pattern:
	print("Searching for %s in %s" % (p, str_line))

	if re.search(p, str_line):
		print("Match found")
	else:
		print("No match found")


# AI Description:
# This code imports the 're' module and defines a list of patterns to search for in a string. It then iterates over each pattern, prints a message indicating the pattern being searched for, and checks if the pattern is found in the given string using regular expressions. Finally, it prints whether a match was found or not for each pattern.

# --- re_sub.py ---
import re

str_line = 'Peter Piper picked a peck of pickled peppers. How many pickled peppers did Peter Piper pick?'
print("Original: ", str_line)

p = re.sub('Peter', 'Mary', str_line)
print("Replaced: ", p)

p = re.sub('Peter', 'Mary', str_line, count=1)
print("Replacing only one occurrence of Peter... ")
print("Replaced: ", p)


# AI Description:
# This code uses the re module to perform text replacement in a string. It replaces all occurrences of 'Peter' with 'Mary' in the original string and then only replaces the first occurrence. The output is printed for both cases.

# --- re_subn.py ---
import re

print("str1:- ")
str1 = "Sky is blue. Sky is beautiful."
print("Original: ", str1)
p = re.subn('beautiful', 'stunning', str1)
print("Replaced: ", p)

print()
print("str_line:- ")
str_line = 'Peter Piper picked a peck of pickled peppers. How many pickled peppers did Peter Piper pick?'
print("Original: ", str_line)
p = re.subn('Peter', 'Mary', str_line)
print("Replaced: ", p)


# AI Description:
# This code uses the re module to perform string substitution in two sentences. The first sentence substitutes "beautiful" with "stunning" in a given string. The second sentence replaces "Peter" with "Mary" in another string and displays the results using the subn function.

# --- shorten_example.py ---
import textwrap

str1 = "Python is an interpreted high-level programming language for general-purpose programming. Created by Guido van Rossum and first released in 1991, \n\nPython has a design philosophy that emphasizes code readability, notably using significant whitespace."

s = textwrap.shorten(str1, width=50)
print(s)


# AI Description:
# The code imports the textwrap module and defines a string 'str1'. The code then shortens 'str1' to fit within a specified width of 50 characters and prints the shortened version.

# --- wrap_example.py ---
import textwrap

sample_string = '''Python is an interpreted high-level programming language for 
general-purpose programming. Created by Guido van Rossum and first released in 			1991, Python has a design philosophy that emphasizes code readability, 
notably using significant whitespace.'''

w = textwrap.wrap(text=sample_string, width=30) 
print(w)


# AI Description:
# This code imports the textwrap module and defines a sample string. It then wraps the sample string to a width of 30 characters per line using the textwrap.wrap() function and prints the resulting list of wrapped lines.


# --- /home/student/GithubApps/Mastering-Python-Scripting-for-System-Administrators-/Chapter08 ---

# --- format_example.py ---
# Using single formatter
print("{}, My name is John".format("Hi"))

str1 = "This is John. I am learning {} scripting language."
print(str1.format("Python"))

print("Hi, My name is Sara and I am {} years old !!".format(26))

# Using multiple formatters
str2 = "This is Mary {}. I work at {} Resource department. I am {} years old !!"
print(str2.format("Jacobs", "Human", 30))

print("Hello {}, Nice to meet you. I am {}.".format("Emily", "Jennifer"))


# AI Description:
# The code uses the `format()` method to insert values into strings. Single formatters are used with curly braces `{}` to replace placeholders with values. Multiple formatters are used to insert multiple values into a single string by passing them as arguments to the `format()` method.

# --- input_example.py ---
str1 = input("Enter a string: ")
print("Entered string is: ", str1)

print()

a = int(input("Enter the value of a: "))
b = int(input("Enter the value of b: "))
c = a + b
print("Value of c is: ", c)

print()

num1 = float(input("Enter num 1: "))
num2 = float(input("Enter num 2: "))
num3 = num1/num2
print("Value of num 3 is: ", num3)



# AI Description:
# The code prompts the user to enter a string, integer values for 'a' and 'b', and floating-point numbers for 'num1' and 'num2'. It then performs calculations based on the inputs and displays the results along with corresponding messages.

# --- print_example.py ---
# printing a simple string on the screen.
print("Hello Python")

# Accessing only a value.
a = 80
print(a)

# printing a string on screen as well as accessing a value.
a = 50
b = 30
c = a/b
print("The value of c is: ", c)


# AI Description:
# The code prints "Hello Python" to the screen and then prints the value of variable 'a', which is 80. It also calculates the value of 'c' as 'a' divided by 'b' and prints "The value of c is: " followed by the result.

# --- send_email.py ---
import smtplib
from email.mime.text import MIMEText
import getpass

host_name = 'smtp.gmail.com'
port = 465
u_name = 'username/emailid'
password = getpass.getpass()
sender = 'sender_name'
receivers = ['receiver1_email_address', 'receiver2_email_address']

text = MIMEText('Test mail')
text['Subject'] = 'Test'
text['From'] = sender
text['To'] = ', '.join(receivers)

s_obj = smtplib.SMTP_SSL(host_name, port)
s_obj.login(u_name, password)
s_obj.sendmail(sender, receivers, text.as_string())
s_obj.quit()
print("Mail sent successfully")



# AI Description:
# This code sends an email using Gmail's SMTP server. It prompts the user for their email password, sets the sender and receiver email addresses, creates a message with a subject and body, sends the email, and prints a success message if the email is sent successfully.

# --- send_email_attachment.py ---
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
import getpass

host_name = 'smtp.gmail.com'  # smtp.mail.yahoo.com
port = 465
u_name = 'mansi.joshi990@gmail.com'
password = getpass.getpass()
sender = 'Mansi Joshi'
receivers = ['info@levanatech.com', 'kalpesh7402patil@gmail.com']

text = MIMEMultipart()
text['Subject'] = 'Test Attachment'
text['From'] = sender
text['To'] = ', '.join(receivers)

txt = MIMEText('Sending a sample image.')
text.attach(txt)

f_path = '/home/student/Desktop/mountain.jpg'
with open(f_path, 'rb') as f:
    img = MIMEImage(f.read())

img.add_header('Content-Disposition',
               'attachment',
               filename=os.path.basename(f_path))
text.attach(img)

server = smtplib.SMTP_SSL(host_name, port)
print("Attachment sent successfully !!")
server.login(u_name, password)
server.sendmail(sender, receivers, text.as_string())
server.quit()


# AI Description:
# This Python code sets up an email with an attachment and sends it using SMTP. It imports necessary modules, sets up email content, attaches an image, and sends the email with the attachment using a Gmail SMTP server. The user is prompted for a password using getpass for security.

# --- stdin_stdout_example.py ---
import sys

print("Enter number1: ")
a = int(sys.stdin.readline())

print("Enter number2: ")
b = int(sys.stdin.readline())

c = a + b
sys.stdout.write("Result: %d " % c)



# AI Description:
# This Python code reads two numbers from the user using stdin and calculates their sum. Finally, it prints the result using stdout in the format "Result: {sum}".

# --- string_formatting.py ---
# Basic formatting
a = 10
b = 30
print("The values of a and b are %d %d" % (a, b))
c = a + b
print("The value of c is %d" % c)

str1 = 'John'
print("My name is %s" % str1)

x = 10.5
y = 33.5
z = x * y
print("The value of z is %f" % z)
print()

# aligning
name = 'Mary'
print("Normal: Hello, I am %s !!" % name)
print("Right aligned: Hello, I am %10s !!" % name)
print("Left aligned: Hello, I am %-10s !!" % name)
print()

# truncating
print("The truncated string is %.4s" % ('Examination'))
print()

# formatting placeholders
students = {'Name' : 'John', 'Address' : 'New York'}
print("Student details: Name:%(Name)s Address:%(Address)s" % students)



# AI Description:
# The code demonstrates basic formatting in Python using different placeholders like %d, %s, %f for integers, strings, and floats respectively. It shows how to align text using formatting options like right-aligned and left-aligned. Additionally, it includes truncating strings and formatting placeholders using dictionaries.


# --- /home/student/GithubApps/Mastering-Python-Scripting-for-System-Administrators-/Chapter09 ---

# --- append_values.py ---
from openpyxl import Workbook

book_obj = Workbook()
excel_sheet = book_obj.active

rows = (
    (11, 12, 13),
    (21, 22, 23),
    (31, 32, 33),
    (41, 42, 43)
)

for values in rows:
	excel_sheet.append(values)
	print()

print("values are successfully appended")
book_obj.save('test.xlsx')


# AI Description:
# This code uses the openpyxl library to create an Excel workbook. It creates a sheet and appends rows of values to it. After appending the values, it saves the workbook to a file named 'test.xlsx'.

# --- create_excel.py ---
from openpyxl import Workbook

book_obj = Workbook()

excel_sheet = book_obj.active
excel_sheet['A1'] = 'Name'
excel_sheet['A2'] = 'student'
excel_sheet['B1'] = 'age'
excel_sheet['B2'] = '24'

book_obj.save("test.xlsx")
print("Excel created successfully")



# AI Description:
# This code creates a new Excel workbook using the openpyxl library. It adds data to the workbook, saves it as "test.xlsx", and prints a success message.

# --- csv_read.py ---
import csv

csv_file = open('test.csv', 'r')
with csv_file:
read_csv = csv.reader(csv_file)
    	for row in read_csv:
      		print(row)



# AI Description:
# This code reads and prints the contents of a CSV file named 'test.csv'. It uses the csv module to handle reading the file and printing each row. The file is opened in read mode and then each row is printed to the console.

# --- csv_write.py ---
import csv

write_csv = [['Name', 'Sport'], ['Andres Iniesta', 'Football'], ['AB de Villiers', 'Cricket'], ['Virat Kohli', 'Cricket'], ['Lionel Messi', 'Football']]

with open('csv_write.csv', 'w') as csvFile:
writer = csv.writer(csvFile)
    	writer.writerows(write_csv)
	print(write_csv)



# AI Description:
# This code creates a list of lists containing names and sports. It then writes this data to a CSV file named 'csv_write.csv' using the csv module. The list of lists is printed to the console after writing to the CSV file.

# --- extract_column_names.py ---
import xlrd  

excel_file = ("/home/student/work/sample.xlsx")

book_obj = xlrd.open_workbook(excel_file) 
excel_sheet = book_obj.sheet_by_index(0) 
excel_sheet.cell_value(0, 0)

for i in range(excel_sheet.ncols): 
	print(excel_sheet.cell_value(0, i))



# AI Description:
# This code imports the xlrd module to work with Excel files. It opens a specific Excel file and reads the first row values. Then, it prints out the values in the first row of the Excel sheet column by column.

# --- extract_text.py ---
import PyPDF2

pdf = open('test.pdf', 'rb') 
read_pdf = PyPDF2.PdfFileReader(pdf) 
pdf_page = read_pdf.getPage(1)
pdf_content = pdf_page.extractText()
print(pdf_content)
pdf.close()



# AI Description:
# This code uses PyPDF2 to read a PDF file named 'test.pdf', extracts the text content of the second page, and then prints it to the console. Finally, it closes the PDF file.

# --- read_excel.py ---
import xlrd 

excel_file = ("/home/student/sample.xlsx") 

book_obj = xlrd.open_workbook(excel_file) 
excel_sheet = book_obj.sheet_by_index(0) 

result = excel_sheet.cell_value(0, 1) 
print(result)


# AI Description:
# This code imports the xlrd module to work with Excel files. It opens a specific Excel file, reads the first sheet, extracts the value in the cell at row 0 and column 1, and then prints that value.

# --- read_multiple.py ---
import openpyxl

book_obj = openpyxl.load_workbook('sample.xlsx')

excel_sheet = book_obj.active

cells = excel_sheet['A1': 'C6']

for c1, c2, c3 in cells:
	print("{0:6} {1:6} {2:6}".format(c1.value, c2.value, c3.value))



# AI Description:
# This code uses the openpyxl library to load an Excel file named 'sample.xlsx' and access its active sheet. It then retrieves cells A1 to C6 from the sheet and prints the values of those cells in a formatted way.

# --- read_pdf.py ---
import PyPDF2

pdf = open('test.pdf', 'rb')
read_pdf = PyPDF2.PdfFileReader(pdf)

print("Number of pages in pdf : ", read_pdf.numPages)
pdf.close()



# AI Description:
# This code uses PyPDF2 library to open and read a PDF file named 'test.pdf'. It then prints the number of pages in the PDF file before closing it.

# --- rotate_pdf.py ---
import PyPDF2
 
pdf = open('test.pdf', 'rb')
rd_pdf = PyPDF2.PdfFileReader(pdf)
wr_pdf = PyPDF2.PdfFileWriter()
 
for pg_num in range(rd_pdf.numPages):
	pdf_page = rd_pdf.getPage(pg_num)
	pdf_page.rotateClockwise(90)
	wr_pdf.addPage(pdf_page)
 
pdf_out = open('rotated.pdf', 'wb')
wr_pdf.write(pdf_out)
pdf_out.close()
print("pdf successfully rotated")
pdf.close()



# AI Description:
# This code uses PyPDF2 library to rotate pages in a PDF file by 90 degrees clockwise. It opens a PDF file, reads its pages, rotates them, and saves the rotated pages to a new PDF file named 'rotated.pdf'. Finally, it prints a success message after rotating the PDF and closes all file objects.

# --- text_read.py ---
text_file = open("test.txt", "r")
data = text_file.read()
print(data)
text_file.close()



# AI Description:
# This code opens a file named "test.txt" in read mode, reads its contents, and then prints the data. Finally, it closes the file.

# --- text_write.py ---
text_file = open("test.txt", "w")
text_file.write("Monday\nTuesday\nWednesday\nThursday\nFriday\nSaturday\n")
text_file.close()



# AI Description:
# This code opens a text file named "test.txt" in write mode, writes the days of the week to the file, and then closes the file.


# --- /home/student/GithubApps/Mastering-Python-Scripting-for-System-Administrators-/Chapter10 ---

# --- client.py ---
import socket

host = socket.gethostname()  # as both code is running on same pc
port = 5000  # socket server port number

client_socket = socket.socket()  # instantiate
client_socket.connect((host, port))  # connect to the server

message = input(" -> ")  # take input

while message.lower().strip() != 'bye':
	client_socket.send(message.encode())  # send message
	data = client_socket.recv(1024).decode()  # receive response

	print('Received from server: ' + data)  # show in terminal

	message = input(" -> ")  # again take input

client_socket.close()  # close the connection


# AI Description:
# This code creates a client socket that connects to a server socket using the host and port number. It sends user input to the server and displays the response, repeating this process until the user enters 'bye'. Finally, the client socket is closed to end the connection.

# --- get_example.py ---
import http.client

con_obj = http.client.HTTPSConnection("www.imdb.com")
con_obj.request("GET", "/")
response = con_obj.getresponse()
print("Status: {}".format(response.status))

read_data = response.read(1000)
print(read_data)
con_obj.close()


# AI Description:
# This Python code establishes a HTTPS connection to www.imdb.com, sends a GET request, and retrieves the response status. It then reads and prints the first 1000 bytes of the response data before closing the connection.

# --- get_ftp_files.py ---
from ftplib import FTP

ftp = FTP('192.168.2.105')
ftp.login('student','training')

ftp.cwd('/home/student/work/')
files = ftp.nlst()

# Print out the files
for file in files:
	print("Downloading..." + file)
	ftp.retrbinary("RETR " + file ,open("/home/student/testing/" + file, 'wb').write)

ftp.close()



# AI Description:
# This code connects to an FTP server at IP address 192.168.2.105 using credentials 'student' and 'training'. It changes the current working directory to '/home/student/work/' and lists the files in that directory. It then downloads each file to the local directory '/home/student/testing/'.

# --- get_header_list.py ---
import http.client

con_obj = http.client.HTTPSConnection("www.imdb.com")
con_obj.request("GET", "/")

response = con_obj.getresponse()
headers_list = response.getheaders()

print("Headers: {}".format(headers_list))


# AI Description:
# This code establishes an HTTPS connection to "www.imdb.com", sends a GET request, and retrieves the response headers. It then prints out the list of headers.

# --- get_welcome_msg.py ---
from ftplib import FTP

ftp = FTP('192.168.2.105')
ftp.login('student','training')

welcome_msg = ftp.getwelcome()
print(welcome_msg)

ftp.close()



# AI Description:
# This code establishes an FTP connection to the server at IP address '192.168.2.105' and logs in using the credentials 'student' and 'training'. It retrieves and prints the welcome message from the FTP server, then closes the connection.

# --- make_connection.py ---
import http.client

con_obj = http.client.HTTPConnection('www.levanatech.com', 80, timeout=100)
print(con_obj)


# AI Description:
# This code imports the http.client module and creates an HTTP connection object to 'www.levanatech.com' on port 80 with a timeout of 100 seconds. The connection object is then printed to the console.

# --- post_example.py ---
import http.client
import json

con_obj = http.client.HTTPSConnection('www.httpbin.org')

headers_list = {'Content-type': 'application/json'}

post_text = {'text': 'Hello World !!'}
json_data = json.dumps(post_text)

con_obj.request('POST', '/post', json_data, headers_list)

response = con_obj.getresponse()
print(response.read().decode())


# AI Description:
# This code sends a POST request to 'www.httpbin.org' with a JSON body containing the text 'Hello World !!'. It then reads and prints the response from the server.

# --- send_command.py ---
from ftplib import FTP

ftp = FTP('192.168.2.105')
ftp.login('student','training')

ftp.cwd('/home/student/')
s_cmd_stat = ftp.sendcmd('STAT')
print(s_cmd_stat)
print()
s_cmd_pwd = ftp.sendcmd('PWD')
print(s_cmd_pwd)
print()

ftp.close()



# AI Description:
# This code connects to an FTP server at '192.168.2.105' with login credentials 'student' and 'training'. It then changes the current working directory to '/home/student/', sends the 'STAT' and 'PWD' commands, and prints the responses. Finally, it closes the FTP connection.

# --- server.py ---
import socket

# get the hostname
host = socket.gethostname()
port = 5000  # initiate port no above 1024

server_socket = socket.socket()  # get instance
# look closely. The bind() function takes tuple as argument
server_socket.bind((host, port))  # bind host address and port together

# configure how many client the server can listen simultaneously
server_socket.listen(2)
conn, address = server_socket.accept()  # accept new connection
print("Connection from: " + str(address))
while True:
	# receive data stream. it won't accept data packet greater than 1024 bytes
	data = conn.recv(1024).decode()
	if not data:
		# if data is not received break
		break
	print("from connected user: " + str(data))
	data = input(' -> ')
	conn.send(data.encode())  # send data to the client

conn.close()  # close the connection



# AI Description:
# This Python code sets up a server that listens for connections on a specified port. When a client connects, it receives and sends data back and forth. The connection is closed when the communication is done.

# --- url_requests_example.py ---
import urllib.request
x = urllib.request.urlopen('https://www.imdb.com/')
print(x.read())



# AI Description:
# This code imports the urllib.request module to fetch data from a URL. It then opens a connection to the IMDb website and reads the content of the webpage, which is then printed to the console.

# --- url_response_header.py ---
import urllib.request

x = urllib.request.urlopen('https://www.imdb.com/')
print(x.info())



# AI Description:
# This code imports the urllib.request module and uses it to open the URL 'https://www.imdb.com/'. It then prints the metadata information of the response obtained from the URL.


# --- /home/student/GithubApps/Mastering-Python-Scripting-for-System-Administrators-/Chapter11 ---

# --- add_attachment.py ---
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
import getpass

host_name = 'smtp.gmail.com'
port = 465

sender = 'mansi.joshi990@gmail.com'
password = getpass.getpass()
receiver = 'kalpesh7402patil@gmail.com'

text = MIMEMultipart()
text['Subject'] = 'Test Attachment'
text['From'] = sender
text['To'] = receiver

txt = MIMEText('Sending a sample image.')
text.attach(txt)

f_path = '/home/student/Desktop/mountain.jpg'
with open(f_path, 'rb') as f:
	img = MIMEImage(f.read())

img.add_header('Content-Disposition',
               'attachment',
               filename=os.path.basename(f_path))
text.attach(img)

s = smtplib.SMTP_SSL(host_name, port)
print("Attachment sent successfully !!")
s.login(sender, password)
s.sendmail(sender, receiver, text.as_string())
s.quit()


# AI Description:
# This Python code sends an email with an attachment using Gmail SMTP. It imports necessary modules, sets up email content and attachment, and sends the email securely. The attachment is a sample image ('mountain.jpg') attached to the email being sent to 'kalpesh7402patil@gmail.com' from 'mansi.joshi990@gmail.com'.

# --- add_html_content.py ---
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import getpass

host_name = 'smtp.gmail.com'
port = 465

sender = 'mansi.joshi990@gmail.com'
password = getpass.getpass()
receiver = 'kalpesh7402patil@gmail.com'

text = MIMEMultipart()
text['Subject'] = 'Test HTML Content'
text['From'] = sender
text['To'] = receiver

msg = """\
<html>
  <body>
    <p>Hello there, <br>
       Good day !!<br>
       <a href="http://www.imdb.com">Home</a> 
    </p>
  </body>
</html>
"""
html_content = MIMEText(msg, "html")
text.attach(html_content)

s = smtplib.SMTP_SSL(host_name, port)
print("Attachment sent successfully !!")
s.login(sender, password)
s.sendmail(sender, receiver, text.as_string())
s.quit()


# AI Description:
# This Python code sends an HTML email using Gmail's SMTP server. It imports necessary modules, sets up email content, and sends the email securely with the SMTP_SSL method. The email includes a simple HTML message with a link. The user is prompted to enter their Gmail password using getpass to authenticate the sender.

# --- all_emails.py ---
import poplib

pop3_server = 'pop.gmail.com'
username = 'Emaild_address'
password = getpass.getpass()

email_obj = poplib.POP3_SSL(pop3_server)
print(email_obj.getwelcome())
email_obj.user(username)
email_obj.pass_(password)
email_stat = email_obj.stat()
NumofMsgs = email_stat[0]
for i in range(NumofMsgs):
	for mail in email_obj.retr(i+1)[1]:
		print(mail)



# AI Description:
# This code imports the poplib module and connects to a Gmail POP3 server using SSL. It retrieves and prints all emails in the inbox using a loop. It prompts the user to enter their email address and password securely.

# --- imap_email.py ---
import imaplib
import pprint

imap_server = 'imap.gmail.com'
username = 'Emaild_address'
password = getpass.getpass()

imap_obj = imaplib.IMAP4_SSL(imap_server)
imap_obj.login(username, password)
imap_obj.select('Inbox')
temp, data_obj = imap_obj.search(None, 'ALL')
for data in data_obj[0].split():
	temp, data_obj = imap_obj.fetch(data, '(RFC822)')
	print('Message: {0}\n'.format(data))
	pprint.pprint(data_obj[0][1])
	break
imap_obj.close()



# AI Description:
# This code connects to a Gmail IMAP server and retrieves all messages from the Inbox. It fetches and prints the content of each message using the pprint module. The loop breaks after processing the first message.

# --- latest_email.py ---
import poplib

pop3_server = 'pop.gmail.com'
username = 'Emaild_address'
password = getpass.getpass()

email_obj = poplib.POP3_SSL(pop3_server)
print(email_obj.getwelcome())
email_obj.user(username)
email_obj.pass_(password)
print("\nLatest Mail\n")
latest_email = email_obj.retr(1)
print(latest_email[1])



# AI Description:
# This code connects to a Gmail POP3 server using SSL, retrieves the latest email, and prints its content. The user is prompted to enter their email address and password securely using the getpass library.

# --- number_of_emails.py ---
import poplib
import getpass

pop3_server = 'pop.gmail.com'
username = 'Emaild_address'
password = getpass.getpass()

email_obj = poplib.POP3_SSL(pop3_server)
print(email_obj.getwelcome())
email_obj.user(username)
email_obj.pass_(password)
email_stat = email_obj.stat()

print("New arrived e-Mails are : %s (%s bytes)" % email_stat)



# AI Description:
# This Python code uses the `poplib` library to connect to a Gmail account via POP3 protocol and retrieve email statistics. The user is prompted to enter their email address and password securely using `getpass.getpass()`. The code then prints the number of new emails and their total size in bytes.

# --- write_email_message.py ---
import smtplib
import getpass

host_name = "smtp.gmail.com"
port = 465

sender = 'mansi.joshi990@gmail.com'
receiver = 'kalpesh7402patil@gmail.com'

password = getpass.getpass()
msg = """\
Subject: Test Mail
Hello from Sender !!"""

s = smtplib.SMTP_SSL(host_name, port)
s.login(sender, password)
s.sendmail(sender, receiver, msg)
s.quit()

print("Mail sent successfully")


# AI Description:
# This code sends a test email using Gmail's SMTP server. It prompts the user for their email password using `getpass.getpass()`, sets up the email message, connects to the SMTP server, logs in with the sender's email and password, sends the email to the receiver, and then prints a success message.


# --- /home/student/GithubApps/Mastering-Python-Scripting-for-System-Administrators-/Chapter12 ---

# --- fabfile.py ---
from fabric.api import *

env.hosts=["host_name@host_ip"]
env.password='your password'

def dir():
    run('mkdir fabric')
    print('Directory named fabric has been created on your host network')

def diskspace():
    run('df')

def check():
    host_type()



# AI Description:
# This code uses Fabric to define three functions: "dir" creates a directory named "fabric" on a specified host, "diskspace" runs the "df" command on the host to check disk space, and "check" calls another function "host_type." The code sets the host and password for the Fabric connection.

# --- nmiko.py ---
from netmiko import ConnectHandler

remote_device={
    'device_type': 'cisco_ios',
    'ip':  'your remote_device ip address',
    'username': 'username',
    'password': 'password',
}

remote_connection = ConnectHandler(**remote_device)
#net_connect.find_prompt()

for n in range (2,6):
      print("Creating VLAN " + str(n))
      commands = ['exit','vlan database','vlan ' + str(n), 'exit']
      output = remote_connection.send_config_set(commands)
      print(output)

command = remote_connection.send_command('show vlan-switch brief')
print(command)



# AI Description:
# This code establishes a connection to a Cisco device using Netmiko, creates VLANs 2 to 5 on the device, displays the output of each VLAN creation command, and then retrieves and prints a summary of VLANs on the device using the 'show vlan-switch brief' command.

# --- pmiko.py ---
import paramiko
import time

ip_address = "host_ip_address"
usr = "host_username"
pwd = "host_password"

c = paramiko.SSHClient()
c.set_missing_host_key_policy(paramiko.AutoAddPolicy())
c.connect(hostname=ip_address,username=usr,password=pwd)

print("SSH connection is successfuly established with ", ip_address)

rc = c.invoke_shell()
for n in range (2,6):
    print("Creating VLAN " + str(n))
    rc.send("vlan database\n")
    rc.send("vlan " + str(n) +  "\n")
    rc.send("exit\n")
    time.sleep(0.5)

time.sleep(1)
output = rc.recv(65535)
print(output)
c.close



# AI Description:
# This code establishes an SSH connection to a host using Paramiko, creates VLANs using a loop, and prints the output received from the host after sending commands.

# --- ssh_using_sub.py ---
import subprocess
import sys

HOST="your host username@host ip"
COMMAND= "ls"

ssh_obj = subprocess.Popen(["ssh", "%s" % HOST, COMMAND],
                       shell=False,
                       stdout=subprocess.PIPE,
                       stderr=subprocess.PIPE)

result = ssh_obj.stdout.readlines()
if result == []:
    err = ssh_obj.stderr.readlines()
    print(sys.stderr, "ERROR: %s" % err)
else:
    print(result)



# AI Description:
# This Python code uses the subprocess module to execute a remote command via SSH. It connects to a specified host and runs the 'ls' command. It then prints the output of the command or any error messages encountered during the execution.

# --- telnet_example.py ---
import telnetlib
import getpass
import sys

HOST_IP = "your host ip address"
host_user = input("Enter your telnet username: ")
password = getpass.getpass()

t = telnetlib.Telnet(HOST_IP)
t.read_until(b"Username:")
t.write(host_user.encode("ascii") + b"\n")
if password:
	t.read_until(b"Password:")
	t.write(password.encode("ascii") + b"\n")

t.write(b"enable\n")
t.write(b"enter_remote_device_password\n") #password of your remote device
t.write(b"conf t\n")
t.write(b"int loop 1\n")
t.write(b"ip add 10.1.1.1 255.255.255.255\n")
t.write(b"int loop 2\n")
t.write(b"ip add 20.2.2.2 255.255.255.255\n")
t.write(b"end\n")
t.write(b"exit\n")
print(t.read_all().decode("ascii") )



# AI Description:
# This code uses the telnetlib library to connect to a host IP address, prompts the user for a username and password using getpass, and then sends commands to configure loopback interfaces on a remote device. Finally, it prints the output received from the telnet session.


# --- /home/student/GithubApps/Mastering-Python-Scripting-for-System-Administrators-/Chapter13 ---

# --- box_layout.py ---
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QVBoxLayout

appl = QApplication([])

make_window = QWidget()
l = QVBoxLayout()

l.addWidget(QPushButton('Button 1'))
l.addWidget(QPushButton('Button 2'))

make_window.setLayout(l)
make_window.show()

appl.exec_()


# AI Description:
# This code creates a PyQt5 application with two buttons displayed vertically in a window. It uses QApplication, QWidget, QPushButton, and QVBoxLayout to achieve this layout. Finally, the application event loop is started with appl.exec_().

# --- print_message.py ---
import sys
from PyQt5.QtWidgets import QApplication, QLabel, QPushButton, QWidget
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtGui import QIcon

class simple_app(QWidget):

	def __init__(self):
		super().__init__()
		self.title = 'Main app window'
		self.left = 20
		self.top = 20
		self.height = 300
		self.width = 400
		self.app_initialize()
		

	def app_initialize(self):
		self.setWindowTitle(self.title)
		self.setGeometry(self.left, self.top, self.height, self.width)
		b = QPushButton('Click', self)
		b.setToolTip('Click on the button !!')
		b.move(100,70)
		self.l = QLabel(self)
		self.l.resize(100,50)
		self.l.move(100,200)
		b.clicked.connect(self.on_click)
		self.show()

	@pyqtSlot()
	def on_click(self):		
		self.l.setText("Hello World")

if __name__ == '__main__':
	appl = QApplication(sys.argv)
	ex = simple_app()
	sys.exit(appl.exec_())




# AI Description:
# This code creates a simple GUI application using PyQt5 with a window, a button, and a label. When the button is clicked, the label displays "Hello World". The application is initialized with the specified window size and title.


# --- /home/student/GithubApps/Mastering-Python-Scripting-for-System-Administrators-/Chapter14 ---

# --- parse_ip_address.py ---
import re
from collections import Counter

r_e = r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1}'

with open("access.log") as f:
	print("Reading Apache log file")
	Apache_log = f.read()
	get_ip = re.findall(r_e,Apache_log)
	no_of_ip = Counter(get_ip)
	for k, v in no_of_ip.items():
		print("Available IP Address in log file " + "=> " + str(k) + " " + "Count "  + "=> " + str(v))



# AI Description:
# This Python code reads an Apache log file, extracts IP addresses using a regular expression, and counts the occurrences of each IP address in the log file. Finally, it prints out the available IP addresses and their counts.

# --- read_apache_log.py ---
def read_apache_log(logfile):
	with open(logfile) as f:
		log_obj = f.read()
		print(log_obj)

if __name__ == '__main__':
	read_apache_log("access.log")



# AI Description:
# This code defines a function `read_apache_log` that reads and prints the content of a given Apache log file. It then calls this function with the file "access.log" if the script is run directly.

# --- simple_log.py ---
f=open('/var/log/kern.log','r')
lines = f.readlines()
for line in lines:
	kern_log = line.split()
	print(kern_log)



# AI Description:
# This code opens the file '/var/log/kern.log' in read mode and reads all its lines. It then splits each line into a list of words using whitespace as the delimiter and prints each resulting list.

# --- simple_log1.py ---
f=open('/var/log/kern.log','r')

lines = f.readlines()
for line in lines:
	kern_log = line.split()[1:3]
	print(kern_log)



# AI Description:
# This code opens the file '/var/log/kern.log' in read mode and reads its content line by line. It then splits each line into a list of words and prints the second and third words from each line.


# --- /home/student/GithubApps/Mastering-Python-Scripting-for-System-Administrators-/Chapter15 ---

# --- json_to_python.py ---
import json

j_obj =  '{ "Name":"Harry", "Age":26, "Department":"HR"}'

p_obj = json.loads(j_obj)

print(p_obj)
print(p_obj["Name"])
print(p_obj["Department"])



# AI Description:
# This code imports the json module and defines a JSON string 'j_obj' with key-value pairs. It then loads the JSON string into a Python object 'p_obj' using json.loads(). Finally, it prints the entire 'p_obj', as well as specific values for the keys "Name" and "Department".

# --- python_object_to_json.py ---
import json

python_dict =  {"Name": "Harry", "Age": 26}
python_list =  ["Mumbai", "Pune"]
python_tuple =  ("Basketball", "Cricket")
python_str =  ("hello_world")
python_int =  (150)
python_float =  (59.66)
python_T =  (True)
python_F =  (False)
python_N =  (None)

json_obj = json.dumps(python_dict)
json_arr1 = json.dumps(python_list)
json_arr2 = json.dumps(python_tuple)
json_str = json.dumps(python_str)
json_num1 = json.dumps(python_int)
json_num2 = json.dumps(python_float)
json_t = json.dumps(python_T)
json_f = json.dumps(python_F)
json_n = json.dumps(python_N)

print("json object : ", json_obj)
print("jason array1 : ", json_arr1)
print("json array2 : ", json_arr2)
print("json string : ", json_str)
print("json number1 : ", json_num1)
print("json number2 : ", json_num2)
print("json true", json_t)
print("json false", json_f)
print("json null", json_n)



# AI Description:
# The code converts Python data types (dictionary, list, tuple, string, int, float, boolean, and None) to JSON format using the json module's dumps() function. It then prints out the JSON representation of each data type.

# --- python_to_json.py ---
import json

emp_dict1 =  '{ "Name":"Harry", "Age":26, "Department":"HR"}'

json_obj = json.dumps(emp_dict1)

print(json_obj)



# AI Description:
# This code imports the json module and creates a dictionary with employee information. The dictionary is then converted to a JSON string using the json.dumps() function and printed to the console.

# --- rest_get_example.py ---
import requests

req = requests.get('https://www.imdb.com/news/top?ref_=nv_tp_nw')
print(req)


# AI Description:
# This Python code uses the requests module to send a GET request to the IMDb news page. The response object is stored in the variable req, and then printed to the console.

# --- rest_post_example.py ---
import requests
import json

url_name = 'http://httpbin.org/post'
data = {"Name" : "John"}
data_json = json.dumps(data)
h = {'Content-type': 'application/json'}

res_obj = requests.post(url_name, data=data_json, headers=h)
print(res_obj)





# AI Description:
# This code sends a POST request to http://httpbin.org/post with JSON data containing the name "John". The response object is printed after making the request.

# --- soap_example.py ---
import zeep

w = 'http://www.soapclient.com/xml/soapresponder.wsdl'
c = zeep.Client(wsdl=w)
print(c.service.Method1('Hello', 'World'))


# AI Description:
# This code imports the Zeep library for SOAP web service communication. It creates a client object using a specific WSDL URL and calls a method named 'Method1' with the arguments 'Hello' and 'World', printing the result.


# --- /home/student/GithubApps/Mastering-Python-Scripting-for-System-Administrators-/Chapter16 ---

# --- extract_from_class.py ---
import requests
from bs4 import BeautifulSoup

page_result = requests.get('https://www.imdb.com/news/top?ref_=nv_nw_tp')
parse_obj = BeautifulSoup(page_result.content, 'html.parser')

top_news = parse_obj.find(class_='news-article__content')

print(top_news)



# AI Description:
# This code uses the requests and BeautifulSoup libraries to scrape the top news article content from IMDb's news page. It sends a GET request to the page and uses BeautifulSoup to parse the HTML content. It then finds and prints the top news article content based on the specified class name.

# --- extract_from_tag.py ---
import requests
from bs4 import BeautifulSoup

page_result = requests.get('https://www.imdb.com/news/top?ref_=nv_nw_tp')
parse_obj = BeautifulSoup(page_result.content, 'html.parser')

top_news = parse_obj.find(class_='news-article__content')
top_news_a_content = top_news.find_all('a')

print(top_news_a_content)



# AI Description:
# This code uses the requests module to fetch a webpage from IMDb's top news section. It then uses BeautifulSoup to parse the HTML content. It finds all the 'a' tags within a specific class in the parsed content and prints them.

# --- extract_from_wikipedia.py ---
import requests
from bs4 import BeautifulSoup

page_result = requests.get('https://en.wikipedia.org/wiki/Portal:History')
parse_obj = BeautifulSoup(page_result.content, 'html.parser')

h_obj = parse_obj.find(class_='hlist noprint')
h_obj_a_content = h_obj.find_all('a')

print(h_obj)
print(h_obj_a_content)



# AI Description:
# This code uses the requests library to retrieve the content of a Wikipedia page about history and then parses it using BeautifulSoup. It finds an HTML element with a specific class and then extracts all the anchor elements within it. Finally, it prints the selected element and its anchor content.

# --- parse_web_page.py ---
import requests
from bs4 import BeautifulSoup

page_result = requests.get('https://www.imdb.com/news/top?ref_=nv_nw_tp')
parse_obj = BeautifulSoup(page_result.content, 'html.parser')

print(parse_obj)



# AI Description:
# This code imports the 'requests' and 'BeautifulSoup' libraries. It makes a request to a webpage and then uses Beautiful Soup to parse the content. Finally, it prints the parsed content of the webpage.


# --- /home/student/GithubApps/Mastering-Python-Scripting-for-System-Administrators-/Chapter17 ---

# --- bar_chart.py ---
import matplotlib.pyplot as plt
from matplotlib import style

style.use('ggplot')

x1 = [4,8,12] 
y1 = [12,16,6]

x2 = [5,9,11]
y2 = [6,16,8]

plt.bar(x1,y1,color = 'g',linewidth=3)
plt.bar(x2,y2,color = 'r',linewidth=3)

plt.title("Bar plot")
plt.xlabel("x axis")
plt.ylabel("y axis")

plt.show()


# AI Description:
# This code creates a bar plot with two sets of data using the 'ggplot' style from matplotlib. The first set is displayed in green bars and the second set in red bars. The plot is labeled with a title, x-axis label, and y-axis label before being displayed.

# --- contour_plotly.py ---
from plotly import tools
import plotly
import plotly.graph_objs as go

trace0 = go.Contour(
    z=[[1, 2, 3, 4, 5, 6, 7, 8],
       [2, 4, 7, 12, 13, 14, 15, 16],
       [3, 1, 6, 11, 12, 13, 16, 17],
       [4, 2, 7, 7, 11, 14, 17, 18],
       [5, 3, 8, 8, 13, 15, 18, 19],
       [7, 4, 10, 9, 16, 18, 20, 19],
       [9, 10, 5, 27, 23, 21, 21, 21]],
     line=dict(smoothing=0),
)

trace1 = go.Contour(
    z=[[1, 2, 3, 4, 5, 6, 7, 8],
       [2, 4, 7, 12, 13, 14, 15, 16],
       [3, 1, 6, 11, 12, 13, 16, 17],
       [4, 2, 7, 7, 11, 14, 17, 18],
       [5, 3, 8, 8, 13, 15, 18, 19],
       [7, 4, 10, 9, 16, 18, 20, 19],
       [9, 10, 5, 27, 23, 21, 21, 21]],
     line=dict(smoothing=0.95),
)

data = tools.make_subplots(rows=1, cols=2,
                          subplot_titles=('Smoothing_not_applied', 'Smoothing_applied'))

data.append_trace(trace0, 1, 1)
data.append_trace(trace1, 1, 2)

plotly.offline.plot(data)



# AI Description:
# This code uses Plotly to create two contour plots with different smoothing levels. The plots are then combined into a subplot with titles. Finally, the combined plot is displayed offline.

# --- expo_array.py ---
import numpy as np

array = np.arange(16)
print("The Array is : ",array)

exp = np.exp(array)
print("exponential of given array is : ", exp) 


# AI Description:
# This code imports the NumPy library, creates an array of integers from 0 to 15, and then calculates the exponential of each element in the array using NumPy's exp function. The code then prints both the original array and the resulting exponential array.

# --- histogram_example.py ---
import matplotlib.pyplot as plt
import numpy as np

x = np.random.randn(500)
plt.hist(x)
plt.show()



# AI Description:
# This code generates 500 random numbers using NumPy, then creates a histogram plot of the data using Matplotlib. Finally, it displays the histogram plot.

# --- line_scatter_plot.py ---
import plotly
import plotly.graph_objs as go
import numpy as np

x_axis = np.linspace(0, 1, 50)
y0_axis = np.random.randn(50)+5
y1_axis = np.random.randn(50)
y2_axis = np.random.randn(50)-5

trace0 = go.Scatter(x = x_axis,y = y0_axis,mode = 'markers',name = 'markers')

trace1 = go.Scatter(x = x_axis,y = y1_axis,mode = 'lines+markers',name = 'lines+markers')

trace2 = go.Scatter(x = x_axis,y = y2_axis,mode = 'lines',name = 'lines')

data_sets = [trace0, trace1, trace2]

plotly.offline.plot(data_sets, filename='line_scatter_plot.html')


# AI Description:
# This code uses Plotly to create a line and scatter plot with three different datasets. It generates random data for three y-axes and plots them against a common x-axis. The resulting plot is saved as 'line_scatter_plot.html'.

# --- mult_dim_array.py ---
import numpy as np

my_list1 = [1,2,3,4]
my_list2 = [11,22,33,44]

my_lists = [my_list1,my_list2]

my_array = np.array(my_lists)

print(my_array)



# AI Description:
# This code imports the NumPy library, creates two lists, combines them into a 2D NumPy array, and then prints the resulting array.

# --- open_image.py ---
import matplotlib.pyplot as plt
import matplotlib.image as mpimg

plt.imshow(mpimg.imread('my_sample_plot1.jpg'))
plt.show()



# AI Description:
# This code uses Matplotlib to display an image file ('my_sample_plot1.jpg'). It reads the image using mpimg.imread() and shows it using plt.imshow() and plt.show().

# --- plotly_box_plot.py ---
import random
import plotly
from numpy import *

N = 50.    

c = ['hsl('+str(h)+',50%'+',50%)' for h in linspace(0, 360, N)]

data_set = [{
    'y': 3.5*sin(pi * i/N) + i/N+(1.5+0.5*cos(pi*i/N))*random.rand(20),
    'type':'box',
    'marker':{'color': c[i]}
    } for i in range(int(N))]

layout = {'xaxis': {'showgrid':False,'zeroline':False, 'tickangle':45,'showticklabels':False},
          'yaxis': {'zeroline':False,'gridcolor':'white'},
          'paper_bgcolor': 'rgb(233,233,233)',
          'plot_bgcolor': 'rgb(233,233,233)',
          }

plotly.offline.plot(data_set)



# AI Description:
# This code generates a box plot with random data colored based on a range of hues. The plot is styled with specific layout settings and displayed using Plotly.

# --- read_csv_dataframe.py ---
import pandas as pd

file_name = 'employee.csv'

df = pd.read_csv(file_name)
print(df)
print('-----------------------------------------')
print(df.head(3))
print('-----------------------------------------')
print(df.tail(1))


# AI Description:
# This code reads data from a CSV file named 'employee.csv' using pandas and stores it in a DataFrame called df. It then prints the entire DataFrame, the first 3 rows, and the last row of the DataFrame.

# --- sample_plotly.py ---
import plotly
from plotly.graph_objs import Scatter, Layout

plotly.offline.plot({
    "data": [Scatter(x=[1, 4, 3, 4], y=[4, 3, 2, 1])],
    "layout": Layout(title="plotly_sample_plot")
})



# AI Description:
# This code uses Plotly to create a scatter plot with x and y data points. The plot has a title "plotly_sample_plot" and is displayed offline using Plotly's offline functionality.

# --- scatter_plot_plotly.py ---
import plotly
import plotly.graph_objs as go
import numpy as np

x_axis = np.random.randn(100)
y_axis = np.random.randn(100)

trace = go.Scatter(x=x_axis, y=y_axis, mode = 'markers')
      
data_set = [trace]

plotly.offline.plot(data_set, filename='scatter_plot')



# AI Description:
# This code uses Plotly to create a scatter plot with 100 random data points on the x and y axes. It generates a scatter plot trace using Plotly's graph objects, then plots the data set offline and saves it as an HTML file named 'scatter_plot'.

# --- scatterplot_example.py ---
import matplotlib.pyplot as plt
import numpy as np

x = np.linspace(-2,2,100)
y = np.random.randn(100)
colors = np.random.rand(100)
plt.scatter(x,y,c=colors)
plt.show()


# AI Description:
# This code generates a scatter plot using Matplotlib with 100 random y-values and corresponding x-values from -2 to 2. Each point is colored randomly. The plot is displayed using plt.show().

# --- series_with_index.py ---
import pandas as pd
import numpy as np

s_data = pd.Series([10, 20, 30, 40], index = ['a', 'b', 'c', 'd'], name = 'numbers')
print(s_data)
print()
print("The data at index 2 is: ", s_data[2])
print("The data from range 1 to 3 are:\n", s_data[1:3])



# AI Description:
# This code creates a Pandas Series named 's_data' with values and corresponding index labels. It then prints the entire series, the data at index 2, and the data within a specific range of indices.

# --- series_without_index.py ---
import pandas as pd
import numpy as np

s_data = pd.Series([10, 20, 30, 40], name = 'numbers')
print(s_data)



# AI Description:
# This code imports the pandas and numpy libraries. It creates a pandas Series named 's_data' with values [10, 20, 30, 40] and a name 'numbers'. Finally, it prints the series 's_data'.

# --- simple_array.py ---
import numpy as np

my_list1 = [1,2,3,4]
my_array1 = np.array(my_list1)
print(my_array1)



# AI Description:
# This code imports the NumPy library, creates a list `my_list1` containing integers, converts the list to a NumPy array `my_array1`, and then prints the array.

# --- simple_plot.py ---
import matplotlib.pyplot as plt
import numpy as np

x = np.linspace(0, 5, 10)
y = x**2
plt.plot(x,y)
plt.title("sample plot")
plt.xlabel("x axis")
plt.ylabel("y axis")
plt.show()



# AI Description:
# This code uses matplotlib and numpy to create a simple plot of y = x^2. It generates 10 points evenly spaced between 0 and 5 along the x-axis, calculates the corresponding y values, and then plots the data with labeled axes and a title. Finally, it displays the plot.

# --- simple_plot2.py ---
import matplotlib.pyplot as plt
from matplotlib import style

style.use('ggplot')

x1 = [0,5,10] 
y1 = [12,16,6]

x2 = [6,9,11]
y2 = [6,16,8]

plt.subplot(2,1,1)
plt.plot(x1, y1, linewidth=3)
plt.title("sample plot")
plt.xlabel("x axis")
plt.ylabel("y axis")

plt.subplot(2,1,2)
plt.plot(x2, y2, color = 'r', linewidth=3)
plt.xlabel("x2 axis")
plt.ylabel("y2 axis")
plt.show()



# AI Description:
# This code imports matplotlib to create plots with a specific style, plots two sets of data in two subplots with different styling, and displays the plots with specified labels and titles.

# --- simple_plot3.py ---
import matplotlib.pyplot as plt
from matplotlib import style

style.use('ggplot')

x1 = [0,5,10] 
y1 = [12,16,6]

x2 = [6,9,11]
y2 = [6,16,8]

plt.figure(1)
plt.plot(x1, y1, color = 'g', linewidth=3)
plt.title("sample plot")
plt.xlabel("x axis")
plt.ylabel("y axis")
plt.savefig('my_sample_plot1.jpg')

plt.figure(2)
plt.plot(x2, y2, color = 'r', linewidth=3)
plt.xlabel("x2 axis")
plt.ylabel("y2 axis")
plt.savefig('my_sample_plot2.jpg')
plt.show()



# AI Description:
# The code uses matplotlib to create two plots. The first plot has green lines connecting points (0, 12), (5, 16), and (10, 6). The second plot has red lines connecting points (6, 6), (9, 16), and (11, 8). Each plot has a title, x-axis label, y-axis label, and is saved as an image file.

# --- size_and_dtype.py ---
import numpy as np

my_list1 = [1,2,3,4]
my_list2 = [11,22,33,44]

my_lists = [my_list1,my_list2]

my_array = np.array(my_lists)

print(my_array)

size = my_array.shape
print(size)

data_type = my_array.dtype
print(data_type)



# AI Description:
# This code imports the NumPy library and creates two lists. It then combines the lists into a NumPy array and prints the array, its shape, and data type.

# --- sqrt_array.py ---
import numpy as np

array = np.arange(16)
print("The Array is : ",array)

Square_root = np.sqrt(array)
print("Square root of given array is : ",Square_root) 


# AI Description:
# This code imports the NumPy library, creates an array with values from 0 to 15, and then calculates the square root of each element in the array using NumPy's sqrt function. The original array and the square root values are then printed to the console.


# --- /home/student/GithubApps/Mastering-Python-Scripting-for-System-Administrators-/Chapter18 ---

